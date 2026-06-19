import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from immo_scanner.models import ScoredProperty

logger = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _auto_width(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)


def _yield_fill(value: float) -> PatternFill:
    if value >= 8:
        return GREEN_FILL
    elif value >= 5:
        return ORANGE_FILL
    return RED_FILL


def export_excel(scored: list[ScoredProperty], output_dir: str, filename: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"{filename}_{timestamp}.xlsx")

    wb = Workbook()

    ws_rank = wb.active
    ws_rank.title = "Classement"
    _write_ranking(ws_rank, scored)

    ws_detail = wb.create_sheet("Details")
    _write_details(ws_detail, scored)

    ws_stats = wb.create_sheet("Statistiques")
    _write_stats(ws_stats, scored)

    wb.save(filepath)
    logger.info(f"Excel exporté : {filepath}")
    return filepath


def _write_ranking(ws, scored: list[ScoredProperty]):
    headers = [
        "Rang", "Score", "Ville", "Code postal", "Type", "Surface (m²)",
        "Prix (€)", "Loyer estimé (€/mois)", "Rendement brut (%)",
        "Rendement net est. (%)", "Prix/m² (€)", "Lien", "Source",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    for i, sp in enumerate(scored, 1):
        p = sp.property
        row = [
            i,
            sp.score,
            p.city,
            p.postal_code,
            p.property_type,
            round(p.surface, 1) if p.surface else "",
            p.price,
            round(sp.monthly_rent, 0) if sp.monthly_rent else "",
            sp.gross_yield,
            sp.net_yield,
            round(p.price_per_sqm, 0) if p.price_per_sqm else "",
            p.url,
            p.source,
        ]
        ws.append(row)
        row_num = i + 1

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).border = BORDER

        yield_cell = ws.cell(row=row_num, column=9)
        if sp.gross_yield > 0:
            yield_cell.fill = _yield_fill(sp.gross_yield)

        net_cell = ws.cell(row=row_num, column=10)
        if sp.net_yield > 0:
            net_cell.fill = _yield_fill(sp.net_yield)

        link_cell = ws.cell(row=row_num, column=12)
        if p.url:
            link_cell.hyperlink = p.url
            link_cell.font = Font(color="0563C1", underline="single")

        ws.cell(row=row_num, column=7).number_format = '#,##0'

    _auto_width(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_details(ws, scored: list[ScoredProperty]):
    headers = [
        "Rang", "Score", "Titre", "Ville", "Code postal", "Adresse",
        "Type", "Pièces", "Surface (m²)", "Étage", "Année construction",
        "DPE", "Prix (€)", "Charges (€)", "Loyer estimé (€/mois)",
        "Loyer/m² (€)", "Rendement brut (%)", "Rendement net (%)",
        "Prix/m² (€)", "Score rendement", "Score prix/m²", "Score tension",
        "Score typologie", "Score surface", "Score fraîcheur",
        "Confiance loyer", "Date annonce", "Latitude", "Longitude",
        "Lien", "Source",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    for i, sp in enumerate(scored, 1):
        p = sp.property
        d = sp.score_details
        row = [
            i, sp.score, p.title[:80] if p.title else "", p.city, p.postal_code,
            p.address, p.property_type, p.rooms or "",
            round(p.surface, 1) if p.surface else "",
            p.floor if p.floor is not None else "",
            p.year_built or "", p.dpe, p.price,
            round(p.charges, 0) if p.charges else "",
            round(sp.monthly_rent, 0) if sp.monthly_rent else "",
            round(sp.rental_estimate.rent_per_sqm, 1) if sp.rental_estimate else "",
            sp.gross_yield, sp.net_yield,
            round(p.price_per_sqm, 0) if p.price_per_sqm else "",
            d.get("yield", ""), d.get("price_sqm", ""), d.get("tension", ""),
            d.get("typology", ""), d.get("surface_fit", ""), d.get("freshness", ""),
            sp.rental_estimate.confidence if sp.rental_estimate else "",
            p.date_posted.strftime("%Y-%m-%d") if p.date_posted else "",
            p.latitude or "", p.longitude or "", p.url, p.source,
        ]
        ws.append(row)

        row_num = i + 1
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).border = BORDER

        link_cell = ws.cell(row=row_num, column=30)
        if p.url:
            link_cell.hyperlink = p.url
            link_cell.font = Font(color="0563C1", underline="single")

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _write_stats(ws, scored: list[ScoredProperty]):
    ws.append(["Statistiques globales"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Métrique", "Valeur"])
    _style_header(ws, 2)
    ws.delete_rows(3)
    ws.insert_rows(3)
    ws.cell(row=3, column=1, value="Métrique").font = HEADER_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    ws.cell(row=3, column=2, value="Valeur").font = HEADER_FONT
    ws.cell(row=3, column=2).fill = HEADER_FILL

    yields = [s.gross_yield for s in scored if s.gross_yield > 0]
    prices = [s.property.price for s in scored if s.property.price > 0]

    stats = [
        ("Nombre total de biens", len(scored)),
        ("Rendement brut moyen", f"{sum(yields)/len(yields):.1f}%" if yields else "N/A"),
        ("Rendement brut médian", f"{sorted(yields)[len(yields)//2]:.1f}%" if yields else "N/A"),
        ("Meilleur rendement", f"{max(yields):.1f}%" if yields else "N/A"),
        ("Prix moyen", f"{sum(prices)//len(prices):,} €" if prices else "N/A"),
        ("Prix médian", f"{sorted(prices)[len(prices)//2]:,} €" if prices else "N/A"),
    ]

    cities: dict[str, int] = {}
    sources: dict[str, int] = {}
    for s in scored:
        c = s.property.city or "Inconnu"
        cities[c] = cities.get(c, 0) + 1
        sources[s.property.source] = sources.get(s.property.source, 0) + 1

    row = 4
    for label, val in stats:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=str(val))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Répartition par ville").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Ville").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Nombre").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    row += 1
    for city, count in sorted(cities.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=city)
        ws.cell(row=row, column=2, value=count)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Répartition par source").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Source").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Nombre").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    row += 1
    for source, count in sorted(sources.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=source)
        ws.cell(row=row, column=2, value=count)
        row += 1

    _auto_width(ws)
